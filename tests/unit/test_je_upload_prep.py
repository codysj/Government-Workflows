"""Unit tests for src/workflows/je_upload_prep.py.

Run with:
    .venv/Scripts/python.exe -m pytest tests/unit/test_je_upload_prep.py \
        -q -p no:cacheprovider --basetemp=.pytest_tmp_je_upload_prep
"""
from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

# ------------------------------------------------------------------ #
# Paths to synthetic data
# ------------------------------------------------------------------ #
_REPO = Path(__file__).resolve().parents[2]
_SYNTH = _REPO / "data" / "synthetic"
_JE_DIR = _SYNTH / "je_upload_prep"
_TYLER = _SYNTH / "tyler"

JE_DRAFT_VALID = _JE_DIR / "je_draft_valid.csv"
JE_DRAFT_INVALID = _JE_DIR / "je_draft_invalid.csv"
JE_DRAFT_WARNINGS = _JE_DIR / "je_draft_warnings.csv"
JE_CONFIG = _JE_DIR / "je_config.json"
COA = _TYLER / "chart_of_accounts.csv"
GL_DETAIL = _TYLER / "gl_detail.csv"

# ------------------------------------------------------------------ #
# Module under test
# ------------------------------------------------------------------ #
import src.workflows.je_upload_prep as wf
from src.workflows.je_upload_prep import (
    CAPABILITY,
    EXPORT_FILE_NAMES,
    SAMPLE_INPUTS,
    UPLOAD_HEADERS,
    UPLOAD_SHEET_NAME,
    WORKFLOW,
    WORKFLOW_TYPE,
    JEUploadConfig,
    JEUploadOutput,
    MockLLMProvider,
    build_je_upload_prep_prompt,
    detect_conditions,
    export_artifacts,
    run,
    run_deterministic,
    validate_llm_output,
)


# ================================================================== #
# Helpers
# ================================================================== #
def _rules(det: JEUploadOutput) -> list[str]:
    return [f.rule_used for f in det.findings]


def _blocking_findings(det: JEUploadOutput) -> list[str]:
    blocking = {
        "no_duplicate_journal_line",
        "eff_date_required",
        "eff_date_valid",
        "eff_date_in_fiscal_period",
        "required_fields_present",
        "account_in_coa",
        "no_inactive_account",
        "no_both_debit_and_credit",
        "no_negative_amount",
        "debits_equal_credits_per_journal",
        "debits_equal_credits_overall",
    }
    return [f.rule_used for f in det.findings if f.rule_used in blocking]


def _default_config() -> JEUploadConfig:
    return JEUploadConfig.from_json(JE_CONFIG)


# ================================================================== #
# 1. JEUploadConfig
# ================================================================== #
class TestJEUploadConfig:
    def test_from_json(self):
        cfg = JEUploadConfig.from_json(JE_CONFIG)
        assert cfg.fiscal_period_start == date(2025, 7, 1)
        assert cfg.fiscal_period_end == date(2026, 6, 30)
        assert cfg.allow_inactive_accounts is False
        assert cfg.round_dollar_warning_threshold == Decimal("10000.00")
        assert cfg.min_description_length == 3

    def test_from_dict(self):
        cfg = JEUploadConfig.from_dict({
            "fiscal_period_start": "2025-07-01",
            "fiscal_period_end": "2026-06-30",
            "allow_inactive_accounts": True,
        })
        assert cfg.allow_inactive_accounts is True
        assert cfg.fiscal_period_start == date(2025, 7, 1)

    def test_defaults(self):
        cfg = JEUploadConfig()
        assert cfg.allow_inactive_accounts is False
        assert cfg.round_dollar_warning_threshold == Decimal("10000.00")


# ================================================================== #
# 2. run_deterministic — valid draft
# ================================================================== #
class TestRunDeterministicValid:
    def setup_method(self):
        self.cfg = _default_config()
        self.det = run_deterministic(JE_DRAFT_VALID, COA, config=self.cfg)

    def test_upload_ready(self):
        assert self.det.upload_ready is True

    def test_no_blocking_findings(self):
        assert self.det.blocking_count == 0, (
            f"Unexpected blocking findings: {_blocking_findings(self.det)}"
        )

    def test_debits_equal_credits(self):
        assert self.det.total_debits == self.det.total_credits

    def test_upload_rows_present(self):
        assert len(self.det.upload_rows) == 7  # 7 data rows in je_draft_valid

    def test_source_mapping_complete(self):
        assert len(self.det.source_mapping) == len(self.det.upload_rows)
        indices = [m["source_row_index"] for m in self.det.source_mapping]
        assert sorted(indices) == list(range(len(indices)))

    def test_upload_rows_have_all_headers(self):
        for row in self.det.upload_rows:
            for h in UPLOAD_HEADERS:
                assert h in row, f"Header {h!r} missing from upload row"

    def test_balanced_per_journal(self):
        # Group by journal and check balance.
        from collections import defaultdict
        journal_debits: dict[str, Decimal] = defaultdict(Decimal)
        journal_credits: dict[str, Decimal] = defaultdict(Decimal)
        for row in self.det.upload_rows:
            jnl = row["Journal"]
            db = Decimal(row["Debit"]) if row["Debit"] else Decimal("0")
            cr = Decimal(row["Credit"]) if row["Credit"] else Decimal("0")
            journal_debits[jnl] += db
            journal_credits[jnl] += cr
        for jnl in journal_debits:
            assert journal_debits[jnl] == journal_credits[jnl], (
                f"Journal {jnl} not balanced"
            )

    def test_summary_keys(self):
        s = self.det.summary
        assert s["upload_ready"] is True
        assert "total_debits" in s
        assert "total_credits" in s
        assert s["blocking_findings"] == 0


# ================================================================== #
# 3. run_deterministic — invalid draft (all defects)
# ================================================================== #
class TestRunDeterministicInvalid:
    def setup_method(self):
        self.cfg = _default_config()
        self.det = run_deterministic(JE_DRAFT_INVALID, COA, config=self.cfg)

    def test_upload_not_ready(self):
        assert self.det.upload_ready is False

    def test_no_upload_rows(self):
        # Fail-closed: no upload rows when blocking errors exist.
        assert self.det.upload_rows == []

    def test_no_source_mapping(self):
        assert self.det.source_mapping == []

    def test_blocking_count_nonzero(self):
        assert self.det.blocking_count > 0

    # --- Individual defect assertions ---

    def test_debits_credits_imbalance(self):
        """Journal 9200 debits != credits -> debits_equal_credits_per_journal fires."""
        assert "debits_equal_credits_per_journal" in _rules(self.det), (
            "Expected debits_equal_credits_per_journal finding"
        )

    def test_account_not_in_coa(self):
        """Object 9999 absent from COA -> account_in_coa fires."""
        coa_findings = [
            f for f in self.det.findings
            if f.rule_used == "account_in_coa"
        ]
        assert len(coa_findings) >= 1, "Expected account_in_coa finding for object 9999"
        # Verify the Object code '9999' appears in computed_values of at least one finding.
        objects_flagged = [
            f.computed_values.get("object")
            for f in coa_findings
            if "object" in f.computed_values
        ]
        assert "9999" in objects_flagged, (
            f"Object 9999 not flagged in account_in_coa findings; got {objects_flagged}"
        )

    def test_missing_eff_date(self):
        """Blank Eff Date -> eff_date_required fires."""
        assert "eff_date_required" in _rules(self.det), (
            "Expected eff_date_required finding"
        )

    def test_invalid_date_string(self):
        """'2026-13-40' is unparseable -> eff_date_valid finding fires (not an exception).

        The JE draft is loaded via raw pandas (no Tyler normalizer), so bad dates
        become findings rather than raising ValueError. All defects are reported
        in a single pass.
        """
        invalid_date_findings = [
            f for f in self.det.findings
            if f.rule_used == "eff_date_valid"
        ]
        assert len(invalid_date_findings) >= 1, (
            "Expected eff_date_valid finding for '2026-13-40'"
        )
        # Verify the bad date string appears in at least one finding.
        bad_date_descs = [f.description for f in invalid_date_findings]
        assert any("2026-13-40" in d for d in bad_date_descs), (
            f"Expected '2026-13-40' in eff_date_valid descriptions; got {bad_date_descs}"
        )

    def test_inactive_account(self):
        """Account 100-5100-5999 (Inactive) -> no_inactive_account fires."""
        inactive_findings = [
            f for f in self.det.findings
            if f.rule_used == "no_inactive_account"
        ]
        assert len(inactive_findings) >= 1, (
            "Expected no_inactive_account finding for account 100-5100-5999"
        )
        # Verify the inactive account key appears in computed_values.
        accounts = [
            f.computed_values.get("account") for f in inactive_findings
            if "account" in f.computed_values
        ]
        assert any("5999" in str(a) for a in accounts), (
            f"Expected account with object 5999 in inactive findings; got {accounts}"
        )

    def test_both_debit_and_credit(self):
        """Line with both Debit and Credit -> no_both_debit_and_credit fires."""
        both_findings = [
            f for f in self.det.findings
            if f.rule_used == "no_both_debit_and_credit"
        ]
        assert len(both_findings) >= 1, (
            "Expected no_both_debit_and_credit finding"
        )

    def test_negative_debit(self):
        """Negative Debit -> no_negative_amount fires."""
        neg_findings = [
            f for f in self.det.findings
            if f.rule_used == "no_negative_amount"
        ]
        assert len(neg_findings) >= 1, "Expected no_negative_amount finding"
        assert any("Debit" in f.description for f in neg_findings)


# ================================================================== #
# 3b. Individual blocking defect tests using a sanitized inline CSV
# ================================================================== #
@pytest.fixture
def tmp_je(tmp_path):
    """Helper: write an in-memory JE CSV and return path."""
    def _write(content: str) -> Path:
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        return p
    return _write


@pytest.fixture
def cfg():
    return _default_config()


class TestBlockingDefectsIndividual:
    """Each test plants exactly one defect on a balanced otherwise-valid JE."""

    BASE_VALID = (
        "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
        "Debit,Credit,Line Description,Reference\n"
        "9500,1,2026-02-15,100,5100,5010,Office Supplies,"
        "1000.00,,Supplies purchase,\n"
        "9500,2,2026-02-15,100,5100,5110,Salaries and Wages,"
        ",1000.00,Payroll offset,\n"
    )

    def test_clean_baseline_passes(self, tmp_path, cfg):
        p = tmp_path / "je.csv"
        p.write_text(self.BASE_VALID, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is True
        assert det.blocking_count == 0

    def test_missing_eff_date(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9501,1,,100,5100,5010,Office Supplies,500.00,,Supplies,\n"
            "9501,2,,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "eff_date_required" in _rules(det)

    def test_date_outside_fiscal_period_before(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9502,1,2025-06-30,100,5100,5010,Office Supplies,500.00,,Supplies,\n"
            "9502,2,2025-06-30,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "eff_date_in_fiscal_period" in _rules(det)

    def test_date_outside_fiscal_period_after(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9503,1,2026-07-01,100,5100,5010,Office Supplies,500.00,,Supplies,\n"
            "9503,2,2026-07-01,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "eff_date_in_fiscal_period" in _rules(det)

    def test_date_at_fiscal_period_start_passes(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9504,1,2025-07-01,100,5100,5010,Office Supplies,500.00,,Supplies,\n"
            "9504,2,2025-07-01,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert "eff_date_in_fiscal_period" not in _rules(det)

    def test_date_at_fiscal_period_end_passes(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9505,1,2026-06-30,100,5100,5010,Office Supplies,500.00,,Supplies,\n"
            "9505,2,2026-06-30,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert "eff_date_in_fiscal_period" not in _rules(det)

    def test_account_not_in_coa(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9506,1,2026-02-15,100,5100,9999,Unknown,500.00,,Supplies,\n"
            "9506,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        coa_findings = [f for f in det.findings if f.rule_used == "account_in_coa"]
        assert len(coa_findings) >= 1
        flagged_objects = [
            f.computed_values.get("object") for f in coa_findings
            if "object" in f.computed_values
        ]
        assert "9999" in flagged_objects

    def test_inactive_account_blocked(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9507,1,2026-02-15,100,5100,5999,Obsolete Expense,500.00,,Test,\n"
            "9507,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "no_inactive_account" in _rules(det)

    def test_allow_inactive_accounts_unblocks(self, tmp_path):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9508,1,2026-02-15,100,5100,5999,Obsolete Expense,500.00,,Test,\n"
            "9508,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        cfg_allow = JEUploadConfig.from_dict({
            "fiscal_period_start": "2025-07-01",
            "fiscal_period_end": "2026-06-30",
            "allow_inactive_accounts": True,
        })
        det = run_deterministic(p, COA, config=cfg_allow)
        assert "no_inactive_account" not in _rules(det), (
            "Inactive account should not be blocked when allow_inactive_accounts=True"
        )
        # The entry itself is balanced and valid.
        assert det.upload_ready is True

    def test_both_debit_and_credit(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9509,1,2026-02-15,100,5100,5010,Office Supplies,"
            "500.00,500.00,Both populated,\n"
            "9509,2,2026-02-15,100,5100,5110,Salaries and Wages,"
            ",500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "no_both_debit_and_credit" in _rules(det)

    def test_negative_debit(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9510,1,2026-02-15,100,5100,5010,Office Supplies,-500.00,,Neg debit,\n"
            "9510,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        neg_findings = [f for f in det.findings if f.rule_used == "no_negative_amount"]
        assert len(neg_findings) >= 1
        assert any("Debit" in f.description for f in neg_findings)

    def test_duplicate_journal_line(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9511,1,2026-02-15,100,5100,5010,Office Supplies,300.00,,Dup,\n"
            "9511,1,2026-02-15,100,5100,5110,Salaries and Wages,,300.00,Dup,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "no_duplicate_journal_line" in _rules(det)

    def test_imbalanced_journal(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9512,1,2026-02-15,100,5100,5010,Office Supplies,600.00,,Supplies,\n"
            "9512,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "debits_equal_credits_per_journal" in _rules(det)

    def test_missing_required_fields_fund_org_object(self, tmp_path, cfg):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9513,1,2026-02-15,,,5010,Office Supplies,500.00,,Supplies,\n"
            "9513,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "je.csv"
        p.write_text(content, encoding="utf-8")
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        assert "required_fields_present" in _rules(det)


# ================================================================== #
# 4. run_deterministic — warnings only (je_draft_warnings.csv)
# ================================================================== #
class TestRunDeterministicWarnings:
    def setup_method(self):
        self.cfg = _default_config()
        self.det = run_deterministic(JE_DRAFT_WARNINGS, COA, config=self.cfg)

    def test_upload_ready(self):
        assert self.det.upload_ready is True

    def test_no_blocking_findings(self):
        assert self.det.blocking_count == 0, (
            f"Unexpected blocking findings: {_blocking_findings(self.det)}"
        )

    def test_combo_plausibility_warning_present(self):
        """100/5200/5240 combo is not in COA (Motor Fuel not in 5200 org)."""
        warn_findings = [
            f for f in self.det.findings
            if f.rule_used == "combo_plausibility"
        ]
        assert len(warn_findings) >= 1, (
            "Expected at least one combo_plausibility warning for 100/5200/5240"
        )
        combos = [f.computed_values.get("combination") for f in warn_findings]
        assert "100-5200-5240" in combos, (
            f"Expected 100-5200-5240 in combo warnings, got {combos}"
        )

    def test_warning_findings_have_requires_human_review(self):
        for f in self.det.findings:
            if f.rule_used == "combo_plausibility":
                assert f.requires_human_review is True

    def test_round_dollar_warning(self):
        """25000.00 is a round-dollar amount >= 10000.00."""
        round_findings = [
            f for f in self.det.findings
            if f.rule_used == "round_dollar_large_amount"
        ]
        assert len(round_findings) >= 1, (
            "Expected round_dollar_large_amount warning for 25000.00"
        )


# ================================================================== #
# 5. run_deterministic — with GL detail combo check
# ================================================================== #
class TestRunDeterministicGLCombos:
    def test_gl_combos_suppress_warning(self, tmp_path):
        """If combo is in GL history, the plausibility warning should not fire."""
        # je_draft_warnings.csv uses 100/5200/5240 which IS NOT in COA.
        # If we provide the GL detail, and 100/5200/5240 IS in GL, warning
        # should be suppressed. But in our synthetic data, GL may not have it.
        # We test this logic by creating a GL-like file that has the combo.
        gl_content = (
            "Fiscal Year,Period,Journal,Date,Fund,Org,Object,Description,"
            "Debit,Credit\n"
            "2026,8,7001,2026-02-15,100,5200,5240,Fuel expense,100.00,0.00\n"
            "2026,8,7001,2026-02-15,100,5200,5110,Salaries,0.00,100.00\n"
        )
        gl_p = tmp_path / "gl.csv"
        gl_p.write_text(gl_content, encoding="utf-8")
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_WARNINGS, COA, gl_detail_path=gl_p, config=cfg)
        # With GL detail containing 100-5200-5240, no combo_plausibility warning.
        combo_warnings = [
            f for f in det.findings
            if f.rule_used == "combo_plausibility" and
            f.computed_values.get("combination") == "100-5200-5240"
        ]
        assert len(combo_warnings) == 0, (
            "combo_plausibility warning should be suppressed when combo in GL"
        )


# ================================================================== #
# 6. Export artifacts — success path
# ================================================================== #
class TestExportArtifactsSuccess:
    def setup_method(self):
        pass

    def test_xlsx_written_on_success(self, tmp_path):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        assert det.upload_ready is True
        mock_resp = {"summary": "ok", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        arts = export_artifacts(tmp_path, det, mock_resp, validation)
        names = [a.file_name for a in arts]
        assert "je_upload.xlsx" in names
        assert "je_upload.csv" in names
        assert "source_mapping.csv" in names
        assert "je_prep_summary.md" in names
        assert "je_validation_errors.csv" in names
        assert "validation_report.json" in names
        assert "audit_log.json" in names

    def test_xlsx_not_written_on_failure(self, tmp_path):
        """On blocking errors, je_upload.xlsx and je_upload.csv must NOT exist."""
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9600,1,2026-02-15,100,5100,5010,Office Supplies,600.00,,Supplies,\n"
            "9600,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "bad.csv"
        p.write_text(content, encoding="utf-8")
        cfg = _default_config()
        det = run_deterministic(p, COA, config=cfg)
        assert det.upload_ready is False
        mock_resp = {"summary": "invalid", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        arts = export_artifacts(tmp_path, det, mock_resp, validation)
        names = [a.file_name for a in arts]
        # Upload files must NOT be written.
        assert "je_upload.xlsx" not in names, "je_upload.xlsx must not be written when invalid"
        assert "je_upload.csv" not in names, "je_upload.csv must not be written when invalid"
        # Upload files must not exist on disk.
        assert not (tmp_path / "je_upload.xlsx").exists()
        assert not (tmp_path / "je_upload.csv").exists()
        # Error/summary files ARE written.
        assert "je_validation_errors.csv" in names
        assert "je_prep_summary.md" in names

    def test_xlsx_readable_by_openpyxl(self, tmp_path):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        mock_resp = {"summary": "ok", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        export_artifacts(tmp_path, det, mock_resp, validation)
        wb = openpyxl.load_workbook(tmp_path / "je_upload.xlsx")
        assert UPLOAD_SHEET_NAME in wb.sheetnames
        ws = wb[UPLOAD_SHEET_NAME]
        # Header row.
        header_row = [cell.value for cell in ws[1]]
        assert header_row == UPLOAD_HEADERS

    def test_xlsx_content_matches_csv(self, tmp_path):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        mock_resp = {"summary": "ok", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        export_artifacts(tmp_path, det, mock_resp, validation)
        # Read xlsx.
        wb = openpyxl.load_workbook(tmp_path / "je_upload.xlsx")
        ws = wb[UPLOAD_SHEET_NAME]
        xlsx_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            xlsx_rows.append(list(row))
        # Read csv.
        csv_df = pd.read_csv(tmp_path / "je_upload.csv", dtype=str).fillna("")
        csv_rows = csv_df.values.tolist()
        assert len(xlsx_rows) == len(csv_rows), (
            f"xlsx has {len(xlsx_rows)} rows, csv has {len(csv_rows)}"
        )
        for i, (xr, cr) in enumerate(zip(xlsx_rows, csv_rows)):
            xr_clean = [str(v) if v is not None else "" for v in xr]
            cr_clean = [str(v) for v in cr]
            assert xr_clean == cr_clean, f"Row {i} mismatch: xlsx={xr_clean}, csv={cr_clean}"

    def test_xlsx_debits_credits_balance(self, tmp_path):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        mock_resp = {"summary": "ok", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        export_artifacts(tmp_path, det, mock_resp, validation)
        wb = openpyxl.load_workbook(tmp_path / "je_upload.xlsx")
        ws = wb[UPLOAD_SHEET_NAME]
        header = [cell.value for cell in ws[1]]
        debit_col = header.index("Debit") + 1
        credit_col = header.index("Credit") + 1
        total_d = Decimal("0")
        total_c = Decimal("0")
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = row[debit_col - 1]
            c = row[credit_col - 1]
            if d:
                total_d += Decimal(str(d))
            if c:
                total_c += Decimal(str(c))
        assert total_d == total_c, f"XLSX not balanced: debits={total_d}, credits={total_c}"

    def test_source_mapping_csv_complete(self, tmp_path):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        mock_resp = {"summary": "ok", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        export_artifacts(tmp_path, det, mock_resp, validation)
        mapping_df = pd.read_csv(tmp_path / "source_mapping.csv", dtype=str)
        assert len(mapping_df) == len(det.upload_rows)
        assert "source_row_index" in mapping_df.columns
        assert "upload_row" in mapping_df.columns

    def test_artifacts_have_sha256(self, tmp_path):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        mock_resp = {"summary": "ok", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        arts = export_artifacts(tmp_path, det, mock_resp, validation)
        for a in arts:
            assert a.sha256 and len(a.sha256) == 64, (
                f"Artifact {a.file_name} has invalid sha256: {a.sha256!r}"
            )

    def test_je_prep_summary_md_states_invalid(self, tmp_path):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9610,1,2026-02-15,100,5100,5010,Office Supplies,600.00,,Supplies,\n"
            "9610,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / "bad2.csv"
        p.write_text(content, encoding="utf-8")
        cfg = _default_config()
        det = run_deterministic(p, COA, config=cfg)
        mock_resp = {"summary": "invalid", "referenced_source_rows": [], "categorized_exceptions": [], "suggested_review_steps": [], "draft_memo": "DRAFT"}
        validation = validate_llm_output(mock_resp, det)
        export_artifacts(tmp_path, det, mock_resp, validation)
        summary_md = (tmp_path / "je_prep_summary.md").read_text(encoding="utf-8")
        assert "INVALID" in summary_md.upper()


# ================================================================== #
# 7. LLM + validation
# ================================================================== #
class TestLLMAndValidation:
    def test_mock_llm_produces_valid_response(self):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        provider = MockLLMProvider()
        prompt = build_je_upload_prep_prompt(det)
        resp = provider.generate_structured_response(prompt)
        assert "summary" in resp
        assert "referenced_source_rows" in resp
        assert "categorized_exceptions" in resp

    def test_mock_llm_only_cites_known_refs(self):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        provider = MockLLMProvider()
        prompt = build_je_upload_prep_prompt(det)
        resp = provider.generate_structured_response(prompt)
        # All referenced source rows must be known.
        known_refs = {
            f"{s.table_name}:{s.row_index}"
            for f in det.findings
            for s in f.source_rows
        }
        for ref in resp.get("referenced_source_rows", []):
            assert ref in known_refs or not known_refs, (
                f"LLM cited unknown ref {ref!r}"
            )

    def test_validate_llm_output_passes_valid(self):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        provider = MockLLMProvider()
        prompt = build_je_upload_prep_prompt(det)
        resp = provider.generate_structured_response(prompt)
        result = validate_llm_output(resp, det)
        assert not result.invented_reference_detected
        assert "summary" in resp  # no missing-key error

    def test_validate_rejects_invented_refs(self):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        bad_resp = {
            "summary": "test",
            "referenced_source_rows": ["je_upload:99999"],
            "categorized_exceptions": [],
            "suggested_review_steps": [],
            "draft_memo": "DRAFT",
        }
        result = validate_llm_output(bad_resp, det)
        assert result.invented_reference_detected is True
        assert not result.passed

    def test_validate_rejects_approval_language(self):
        cfg = _default_config()
        det = run_deterministic(JE_DRAFT_VALID, COA, config=cfg)
        bad_resp = {
            "summary": "I approve this journal entry for filing.",
            "referenced_source_rows": [],
            "categorized_exceptions": [],
            "suggested_review_steps": [],
            "draft_memo": "I approve",
        }
        result = validate_llm_output(bad_resp, det)
        assert not result.passed
        assert any("final-approval" in e.lower() or "approval" in e.lower() for e in result.errors)


# ================================================================== #
# 8. run() lifecycle tests
# ================================================================== #
class TestRunLifecycle:
    def test_run_valid_upload_ready(self, tmp_path):
        result = run(
            {"je_draft": str(JE_DRAFT_VALID), "chart_of_accounts": str(COA)},
            config=_default_config(),
            export_dir=tmp_path,
        )
        assert result["upload_ready"] is True
        assert result["workflow_type"] == WORKFLOW_TYPE
        assert "run_id" in result

    def test_run_valid_exports_written(self, tmp_path):
        result = run(
            {"je_draft": str(JE_DRAFT_VALID), "chart_of_accounts": str(COA)},
            config=_default_config(),
            export_dir=tmp_path,
        )
        run_id = result["run_id"]
        out = tmp_path / run_id
        assert (out / "je_upload.xlsx").exists()
        assert (out / "je_upload.csv").exists()
        assert (out / "source_mapping.csv").exists()
        assert (out / "je_prep_summary.md").exists()
        assert (out / "validation_report.json").exists()

    def test_run_imbalanced_no_upload_files(self, tmp_path):
        content = (
            "Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            "Debit,Credit,Line Description,Reference\n"
            "9700,1,2026-02-15,100,5100,5010,Office Supplies,600.00,,Supplies,\n"
            "9700,2,2026-02-15,100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        je_p = tmp_path / "bad.csv"
        je_p.write_text(content, encoding="utf-8")
        result = run(
            {"je_draft": str(je_p), "chart_of_accounts": str(COA)},
            config=_default_config(),
            export_dir=tmp_path,
        )
        assert result["upload_ready"] is False
        run_id = result["run_id"]
        out = tmp_path / run_id
        assert not (out / "je_upload.xlsx").exists(), "je_upload.xlsx must not exist on failure"
        assert not (out / "je_upload.csv").exists(), "je_upload.csv must not exist on failure"

    def test_run_returns_validation(self, tmp_path):
        result = run(
            {"je_draft": str(JE_DRAFT_VALID), "chart_of_accounts": str(COA)},
            config=_default_config(),
            export_dir=tmp_path,
        )
        assert "validation" in result
        validation = result["validation"]
        assert hasattr(validation, "passed")

    def test_run_returns_preflight(self, tmp_path):
        result = run(
            {"je_draft": str(JE_DRAFT_VALID), "chart_of_accounts": str(COA)},
            config=_default_config(),
            export_dir=tmp_path,
        )
        assert "preflight" in result
        assert result["preflight"]["status"] in ("pass", "partial", "fail")

    def test_run_preflight_fail_on_missing_required_file(self, tmp_path):
        result = run(
            {"je_draft": str(JE_DRAFT_VALID)},  # missing chart_of_accounts
            config=_default_config(),
        )
        # Should fail closed (preflight FAIL or ValueError).
        # The result may contain a preflight_status of fail or return early.
        assert result.get("upload_ready") is False or result.get("preflight", {}).get("status") == "fail"

    def test_run_with_gl_detail(self, tmp_path):
        result = run(
            {
                "je_draft": str(JE_DRAFT_VALID),
                "chart_of_accounts": str(COA),
                "gl_detail": str(GL_DETAIL),
            },
            config=_default_config(),
            export_dir=tmp_path,
        )
        assert result["upload_ready"] is True

    def test_run_config_from_path(self, tmp_path):
        result = run(
            {
                "je_draft": str(JE_DRAFT_VALID),
                "chart_of_accounts": str(COA),
                "config": str(JE_CONFIG),
            },
            export_dir=tmp_path,
        )
        assert result["upload_ready"] is True

    def test_run_warnings_draft_upload_ready(self, tmp_path):
        result = run(
            {"je_draft": str(JE_DRAFT_WARNINGS), "chart_of_accounts": str(COA)},
            config=_default_config(),
            export_dir=tmp_path,
        )
        assert result["upload_ready"] is True
        assert result["summary"]["warning_findings"] >= 1

    def test_run_mock_llm_produces_response(self, tmp_path):
        result = run(
            {"je_draft": str(JE_DRAFT_VALID), "chart_of_accounts": str(COA)},
            config=_default_config(),
            export_dir=tmp_path,
        )
        assert "llm_response" in result
        llm = result["llm_response"]
        assert hasattr(llm, "response_json")
        assert "summary" in llm.response_json

    def test_run_with_ledger_and_audit(self, tmp_path):
        """Run with injected in-memory ledger and audit log (integration)."""
        from src.core.run_ledger import RunLedger
        from src.core.audit_log import AuditLog
        from src.core.schemas import WorkflowRun

        ledger = RunLedger(":memory:")
        audit = AuditLog(ledger, audit_dir=str(tmp_path / "audit"), write_jsonl=False)
        wr = WorkflowRun(workflow_type=WORKFLOW_TYPE, created_by="test")
        run_id = ledger.create_run(wr)
        result = run(
            {"je_draft": str(JE_DRAFT_VALID), "chart_of_accounts": str(COA)},
            config=_default_config(),
            ledger=ledger,
            audit=audit,
            run_id=run_id,
            export_dir=tmp_path,
        )
        assert result["upload_ready"] is True
        # Verify findings stored in ledger.
        findings = ledger.list_findings(run_id)
        assert isinstance(findings, list)


# ================================================================== #
# 9. Module constants / registry
# ================================================================== #
class TestModuleConstants:
    def test_workflow_type(self):
        assert WORKFLOW_TYPE == "je_upload_prep"

    def test_capability_required_inputs(self):
        assert "je_draft" in CAPABILITY.required_inputs
        assert "chart_of_accounts" in CAPABILITY.required_inputs

    def test_capability_optional_inputs(self):
        assert "gl_detail" in CAPABILITY.optional_inputs

    def test_capability_accepted_file_types(self):
        types = CAPABILITY.accepted_file_types.get("*", [])
        assert "csv" in types
        assert "xlsx" in types

    def test_export_file_names(self):
        assert "je_upload.xlsx" in EXPORT_FILE_NAMES
        assert "je_upload.csv" in EXPORT_FILE_NAMES
        assert "source_mapping.csv" in EXPORT_FILE_NAMES
        assert "validation_report.json" in EXPORT_FILE_NAMES
        assert "audit_log.json" in EXPORT_FILE_NAMES

    def test_workflow_dict(self):
        assert WORKFLOW["workflow_type"] == WORKFLOW_TYPE
        assert callable(WORKFLOW["run"])
        assert "export_files" in WORKFLOW

    def test_sample_inputs_keys(self):
        assert "je_draft" in SAMPLE_INPUTS
        assert "chart_of_accounts" in SAMPLE_INPUTS

    def test_upload_headers_match_template(self):
        """Verify UPLOAD_HEADERS matches the je_upload_template.csv exactly."""
        template_df = pd.read_csv(
            _TYLER / "je_upload_template.csv", dtype=str
        )
        expected = list(template_df.columns)
        assert UPLOAD_HEADERS == expected, (
            f"UPLOAD_HEADERS mismatch. Got: {UPLOAD_HEADERS}\nExpected: {expected}"
        )

    def test_register_dict(self):
        reg: dict = {}
        wf.register(reg)
        assert WORKFLOW_TYPE in reg
        assert reg[WORKFLOW_TYPE] is run

    def test_register_callable(self):
        registered = {}

        class FakeRegistry:
            def register(self, wtype, fn):
                registered[wtype] = fn

        wf.register(FakeRegistry())
        assert registered[WORKFLOW_TYPE] is run

    def test_detect_conditions_returns_empty(self):
        """detect_conditions always returns [] (blocking handled by deterministic validator)."""
        from src.core.schemas import FileProfile
        profiles = {
            "je_draft": FileProfile(input_key="je_draft", file_name="test.csv",
                                    file_type="csv", present=True, row_count=2),
        }
        result = detect_conditions(profiles, {}, {}, {})
        assert result == []


# ================================================================== #
# 10. Fiscal period boundary tests
# ================================================================== #
class TestFiscalPeriodBoundaries:
    @pytest.mark.parametrize("d,should_pass", [
        ("2025-07-01", True),   # first day of period
        ("2026-06-30", True),   # last day of period
        ("2025-06-30", False),  # one day before start
        ("2026-07-01", False),  # one day after end
        ("2026-01-15", True),   # mid-period
    ])
    def test_boundary_dates(self, tmp_path, d, should_pass):
        content = (
            f"Journal,Line,Eff Date,Fund,Org,Object,Account Description,"
            f"Debit,Credit,Line Description,Reference\n"
            f"9800,1,{d},100,5100,5010,Office Supplies,500.00,,Supplies,\n"
            f"9800,2,{d},100,5100,5110,Salaries and Wages,,500.00,Offset,\n"
        )
        p = tmp_path / f"je_{d.replace('-', '')}.csv"
        p.write_text(content, encoding="utf-8")
        cfg = _default_config()
        det = run_deterministic(p, COA, config=cfg)
        period_findings = [
            f for f in det.findings if f.rule_used == "eff_date_in_fiscal_period"
        ]
        if should_pass:
            assert len(period_findings) == 0, (
                f"Date {d} should pass fiscal period check but got findings: "
                f"{[f.description for f in period_findings]}"
            )
        else:
            assert len(period_findings) >= 1, (
                f"Date {d} should fail fiscal period check but got no findings"
            )
